"""
Restaura orders, despachos y unloadings a Supabase desde la exportacion
generada por la app (hojas Registros, Despachos, Descargues, Citas, Resumen).

⚠️ ANTES DE USAR:
  1. Exporta los datos desde la app (botón Exportar a Excel). El archivo queda
     en Descargas como pedidos_YYYY-MM-DD.xlsx.
  2. Pasa la ruta del archivo con --file o edita XLSX abajo.
  3. Corre en modo dry-run para ver los totales antes de insertar.

Uso:
  python3 restore_data.py --file /ruta/pedidos_YYYY-MM-DD.xlsx
      # dry-run: no inserta nada, solo reporta totales y operarios
  python3 restore_data.py --file /ruta/pedidos_YYYY-MM-DD.xlsx --commit
      # inserta en Supabase

Dependencias: python3 -m pip install openpyxl requests
Credenciales: se leen de frontend/.env (VITE_SUPABASE_URL y VITE_SUPABASE_ANON_KEY).
"""
import sys
import re
import math
import openpyxl
import requests
from collections import Counter

DRY_RUN = "--commit" not in sys.argv

def arg_value(flag, default=None):
    if flag in sys.argv:
        i = sys.argv.index(flag)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return default

XLSX = arg_value("--file", "/home/oscar/Descargas/pedidos_2026-08-06 (1).xlsx")

# Leer credenciales del .env del frontend
def load_env(path):
    env = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

env = load_env("/home/oscar/pedidos/frontend/.env")
URL = env["VITE_SUPABASE_URL"]
ANON = env["VITE_SUPABASE_ANON_KEY"]

HEADERS = {
    "apikey": ANON,
    "Authorization": f"Bearer {ANON}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

def api(method, table, payload=None, params=None):
    r = requests.request(method, f"{URL}/rest/v1/{table}", headers=HEADERS,
                         json=payload, params=params, timeout=60)
    if r.status_code >= 300:
        raise RuntimeError(f"{table} {method} -> {r.status_code}: {r.text[:300]}")
    return r.json() if r.text else []

def parse_time_spent_to_hours(t):
    if not t:
        return 0
    m = re.match(r"(?:(\d+)h)?\s*(?:(\d+)m)?", str(t))
    h = int(m.group(1)) if m and m.group(1) else 0
    mi = int(m.group(2)) if m and m.group(2) else 0
    return h + mi / 60

def calc_kg_per_hour(kg, hours):
    if not hours:
        return 0
    return round(kg / hours * 100) / 100

def calc_efficiency(kgph):
    return round(kgph / 2500 * 10000) / 100

def parse_percent(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace("%", "").strip())
    except ValueError:
        return None

def parse_iso(ts):
    # "2026-08-06 17:28" -> "2026-08-06T17:28:00"
    if not ts:
        return None
    s = str(ts).strip()
    s = s.replace(" ", "T")
    if len(s) == 16:
        s += ":00"
    return s

wb = openpyxl.load_workbook(XLSX, read_only=True)

# ───────────────────────────────────────────────
# 1. Leer Despachos (tienen el order_id: col Orden)
# ───────────────────────────────────────────────
desp_rows = list(wb["Despachos"].iter_rows(values_only=True))[1:]
despachos = []
for r in desp_rows:
    # 0 Fecha,1 Orden,2 Ruta,3 Placa,4 PLC,5 Kg,6 Hora inicio,7 Hora fin,
    # 8 Tiempo cargue,9 Novedad,10 Cant ref,11 Registrado,12 Por
    if not r[1]:
        continue
    despachos.append({
        "order_id_old": int(r[1]),
        "date": str(r[0]).strip() if r[0] else None,
        "ruta": str(r[2]).strip() if r[2] else "",
        "placa": str(r[3]).strip() if r[3] else "",
        "plc": str(r[4]).strip() if r[4] else "",
        "kg": float(r[5]) if r[5] is not None else 0,
        "cargue_start": str(r[6]).strip() if r[6] else "",
        "cargue_end": str(r[7]).strip() if r[7] else "",
        "cargue_time": str(r[8]).strip() if r[8] else "",
        "novedad": str(r[9]).strip().lower() == "sí",
        "cantidad_referencias_novedad": int(r[10]) if r[10] is not None else 0,
        "created_at": parse_iso(r[11]),
        "created_by": str(r[12]).strip().lower() if r[12] else "",
    })

# ───────────────────────────────────────────────
# 2. Leer Registros (info de cada pedido por despacho)
#    y matchear cada fila con su despacho (y order_id)
# ───────────────────────────────────────────────
reg_rows = list(wb["Registros"].iter_rows(values_only=True))[1:]

def rkey(r):
    return (str(r[11]).strip() if r[11] else "", (r[4] or "").strip().upper(),
            (r[3] or "").strip().upper(), r[13] if r[13] is not None else "")

def dkey(d):
    return (d["created_at"][:10] if d["created_at"] else "", d["placa"].upper(),
            d["plc"].upper(), d["kg"])

# cola de despachos por clave (consumir uno a uno por duplicados)
from collections import defaultdict, deque
desp_by_key = defaultdict(deque)
for d in despachos:
    desp_by_key[dkey(d)].append(d)

# info de pedido por order_id
order_info = {}   # order_id_old -> dict de campos del pedido
rows_sin_despacho = []

for r in reg_rows:
    # 1 Fecha,2 Cliente,5 SKU,6 Tipo,7 Kg,8 Operario,9 Eficiencia,
    # 10 Tiempo alistamiento,14 Dev kg,15 Dev bodega,18 Novedad cargue,19 Cant ref
    has_desp = bool(r[3] or r[4])
    info = {
        "date": str(r[1]).strip() if r[1] else None,
        "cliente": str(r[2]).strip() if r[2] else "",
        "sku": str(r[5]).strip() if r[5] is not None else "",
        "type": str(r[6]).strip() if r[6] else "Masivo",
        "kg": float(r[7]) if r[7] is not None else 0,
        "operator": str(r[8]).strip().lower() if r[8] else "",
        "efficiency_pct": parse_percent(r[9]),
        "time_spent": str(r[10]).strip() if r[10] else None,
        "fecha_despacho": str(r[11]).strip() if r[11] else "",
        "cargue_time": str(r[12]).strip() if r[12] else "",
        "kg_despachado": float(r[13]) if r[13] is not None else 0,
        "devolucion_kg": float(r[14]) if r[14] is not None else 0,
        "notas_devolucion": str(r[15]).strip() if r[15] else "",
        "novedad_cargue": str(r[18]).strip().lower() == "sí" if r[18] else False,
        "cant_novedad": int(r[19]) if r[19] is not None else 0,
    }
    if has_desp:
        key = rkey(r)
        q = desp_by_key.get(key)
        if q:
            d = q.popleft()
            order_info.setdefault(d["order_id_old"], info)
        else:
            print(f"AVISO: Registro sin despacho coincidente: {key}")
    else:
        rows_sin_despacho.append(info)

# ───────────────────────────────────────────────
# 3. Construir orders a insertar
# ───────────────────────────────────────────────
orders = {}  # order_id_old -> order dict

# despachados (con despacho)
desp_by_order = defaultdict(list)
for d in despachos:
    desp_by_order[d["order_id_old"]].append(d)

for oid in sorted(desp_by_order):
    info = order_info.get(oid)
    if not info:
        continue
    ds = desp_by_order[oid]
    kg_total = sum(d["kg"] for d in ds)
    hours = parse_time_spent_to_hours(info["time_spent"])
    kgph = calc_kg_per_hour(kg_total, hours) if hours else 0
    eff = info["efficiency_pct"]
    if eff is None and hours:
        eff = calc_efficiency(kgph)
    orders[oid] = {
        "date": info["date"],
        "cliente": info["cliente"],
        "sku": info["sku"],
        "kg": kg_total,
        "operator": info["operator"],
        "start_time": "",
        "end_time": None,
        "type": "Masivo" if info["type"].lower() == "masivo" else "Venta Directa",
        "status": "despachado",
        "time_spent": info["time_spent"],
        "kg_per_hour": kgph if kgph else None,
        "efficiency": eff,
        "plc": ds[0]["plc"] or None,
        "placa": ds[0]["placa"] or None,
        "cargue_start": ds[0]["cargue_start"] or None,
        "cargue_end": ds[0]["cargue_end"] or None,
        "cargue_time": ds[0]["cargue_time"] or None,
        "despachado_kg": kg_total,
        "devolucion_kg": info["devolucion_kg"],
        "notas_devolucion": info["notas_devolucion"],
        "created_by": ds[0]["created_by"],
        "created_at": ds[0]["created_at"],
    }

# pedidos sin despacho
for info in rows_sin_despacho:
    hours = parse_time_spent_to_hours(info["time_spent"])
    kgph = calc_kg_per_hour(info["kg_despachado"], hours) if hours else 0
    eff = info["efficiency_pct"]
    if eff is None and hours:
        eff = calc_efficiency(kgph)
    orders[("nodisp", id(info))] = {
        "date": info["date"],
        "cliente": info["cliente"],
        "sku": info["sku"],
        "kg": 0,
        "operator": info["operator"],
        "start_time": "",
        "end_time": None,
        "type": "Masivo" if info["type"].lower() == "masivo" else "Venta Directa",
        "status": "completed",
        "time_spent": info["time_spent"],
        "kg_per_hour": kgph if kgph else None,
        "efficiency": eff,
        "plc": None,
        "placa": None,
        "cargue_start": None,
        "cargue_end": None,
        "cargue_time": None,
        "despachado_kg": 0,
        "devolucion_kg": info["devolucion_kg"],
        "notas_devolucion": info["notas_devolucion"],
        "created_by": "",
        "created_at": None,
    }

# ───────────────────────────────────────────────
# 4. Unloadings
# ───────────────────────────────────────────────
unload_rows = list(wb["Descargues"].iter_rows(values_only=True))[1:]
unloadings = []
for r in unload_rows:
    # 0 Fecha,1 PTM,2 Kg,3 Operarios,4 Hora inicio,5 Hora final,6 Tiempo,7 Novedad,8 Resuelta
    unloadings.append({
        "date": str(r[0]).strip() if r[0] else None,
        "ptm": str(r[1]).strip() if r[1] else "",
        "kg": float(r[2]) if r[2] is not None else 0,
        "operators": [o.strip() for o in str(r[3]).split(",") if o.strip()] if r[3] else [],
        "start_time": str(r[4]).strip() if r[4] else "",
        "end_time": str(r[5]).strip() if r[5] else "",
        "time_spent": str(r[6]).strip() if r[6] else None,
        "novedad": str(r[7]).strip() if r[7] else None,
        "novedad_resuelta": str(r[8]).strip().lower() == "sí",
        "created_by": "",
    })

wb.close()

# ───────────────────────────────────────────────
# Reporte
# ───────────────────────────────────────────────
print(f"ORDERS a insertar: {len(orders)}")
print(f"  con despacho: {sum(1 for k in orders if not (isinstance(k, tuple)))}")
print(f"  sin despacho: {len(rows_sin_despacho)}")
print(f"DESPACHOS a insertar: {len(despachos)}")
print(f"UNLOADINGS a insertar: {len(unloadings)}")

operadores = sorted({o["operator"] for o in orders.values() if o["operator"]})
print(f"OPERARIOS usados: {operadores}")

# verificar contra Resumen
total_kg = sum(o["kg"] for o in orders.values())
total_orders = len(orders)
print(f"\nTotales calculados: {total_orders} pedidos, {total_kg:.2f} kg despachados")
print(f"Resumen Excel: 522 pedidos (expandidos), 817332.71 kg")

if DRY_RUN:
    print("\n=== DRY RUN: no se inserta nada. Ejecutar con --commit para restaurar ===")
    sys.exit(0)

# ───────────────────────────────────────────────
# 5. Insertar en Supabase
# ───────────────────────────────────────────────
print("\n=== Insertando orders ===")
order_map = {}  # key -> nuevo id
order_keys = list(orders.keys())
BATCH = 50
for i in range(0, len(order_keys), BATCH):
    batch = order_keys[i:i + BATCH]
    rows = [orders[k] for k in batch]
    inserted = api("POST", "orders", rows)
    print(f"  lote {i//BATCH+1}: +{len(inserted)}")
    for k, row in zip(batch, inserted):
        order_map[str(k)] = row["id"]

print(f"Total orders insertados: {len(order_map)}")

print("\n=== Insertando despachos ===")
desp_insert = []
for d in despachos:
    new_oid = order_map.get(str(d["order_id_old"]))
    if new_oid is None:
        print(f"  OJO: despacho sin order mapeada: {d}")
        continue
    desp_insert.append({
        "order_id": new_oid,
        "ruta": d["ruta"],
        "placa": d["placa"],
        "plc": d["plc"],
        "kg": d["kg"],
        "date": d["date"],
        "cargue_start": d["cargue_start"],
        "cargue_end": d["cargue_end"],
        "cargue_time": d["cargue_time"],
        "novedad": d["novedad"],
        "cantidad_referencias_novedad": d["cantidad_referencias_novedad"],
        "created_by": d["created_by"],
        "created_at": d["created_at"],
    })
for i in range(0, len(desp_insert), BATCH):
    batch = desp_insert[i:i + BATCH]
    inserted = api("POST", "despachos", batch)
    print(f"  lote {i//BATCH+1}: +{len(inserted)}")

print("\n=== Insertando unloadings ===")
for i in range(0, len(unloadings), BATCH):
    batch = unloadings[i:i + BATCH]
    inserted = api("POST", "unloadings", batch)
    print(f"  lote {i//BATCH+1}: +{len(inserted)}")

print("\n=== Listo ===")
